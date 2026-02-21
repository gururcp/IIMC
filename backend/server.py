from fastapi import FastAPI, APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone, date as date_type
from pathlib import Path
from dotenv import load_dotenv
import os, logging, io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ProgressUpdate(BaseModel):
    progress: float
    update_notes: Optional[str] = ""

class DateUpdate(BaseModel):
    start_date: str
    end_date: str
    update_notes: Optional[str] = ""

class RiskUpdate(BaseModel):
    risk_flagged: bool
    risk_notes: Optional[str] = ""
    update_notes: Optional[str] = ""

class NotesUpdate(BaseModel):
    notes: str


async def log_history(task_id: int, action: str, field: str, old_value, new_value, notes: str = ""):
    await db.task_history.insert_one({
        "task_id": task_id,
        "action": action,
        "field": field,
        "old_value": str(old_value),
        "new_value": str(new_value),
        "notes": notes,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })


@app.on_event("startup")
async def startup_event():
    count = await db.tasks.count_documents({})
    if count == 0:
        from seed_data import get_seed_tasks
        tasks = get_seed_tasks()
        await db.tasks.insert_many(tasks)
        logger.info(f"Seeded {len(tasks)} tasks")
    await db.tasks.create_index("task_id", unique=True)
    await db.tasks.create_index("parent_task_id")
    await db.tasks.create_index("phase")
    await db.task_history.create_index("task_id")
    await db.task_history.create_index("timestamp")
    logger.info("ConstructOS API started")


def determine_status(progress, task):
    if task.get("risk_flagged") and progress < 100:
        return "at_risk"
    if progress == 0:
        return "not_started"
    if progress >= 100:
        return "completed"
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if today > task.get("end_date", "9999-12-31"):
        return "delayed"
    return "in_progress"


async def rollup_progress(parent_task_id):
    children = await db.tasks.find({"parent_task_id": parent_task_id}, {"_id": 0}).to_list(None)
    if not children:
        return
    effective = [c for c in children if not c.get("exclude_from_rollup", False)]
    if not effective:
        return
    total_dur = sum(c["duration"] for c in effective)
    if total_dur == 0:
        progress = sum(c["progress"] for c in effective) / len(effective)
    else:
        progress = sum(c["progress"] * c["duration"] for c in effective) / total_dur
    progress = round(min(progress, 100), 1)

    if progress == 0:
        status = "not_started"
    elif progress >= 100:
        status = "completed"
    else:
        status = "in_progress"

    parent = await db.tasks.find_one({"task_id": parent_task_id}, {"_id": 0})
    if not parent:
        return

    has_delayed = any(c["status"] == "delayed" for c in effective)
    has_risk = any(c["status"] == "at_risk" or c.get("risk_flagged") for c in effective)

    if parent.get("risk_flagged") and progress < 100:
        status = "at_risk"
    elif has_delayed and progress < 100:
        status = "delayed"
    elif has_risk and progress < 100:
        status = "at_risk"

    await db.tasks.update_one(
        {"task_id": parent_task_id},
        {"$set": {"progress": progress, "status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if parent.get("parent_task_id") is not None:
        await rollup_progress(parent["parent_task_id"])


@api_router.get("/")
async def root():
    return {"message": "ConstructOS API - IIMC Amravati Project Tracker"}


@api_router.get("/tasks")
async def get_tasks(phase: Optional[str] = None):
    query = {}
    if phase and phase != "all":
        query["phase"] = phase
    tasks = await db.tasks.find(query, {"_id": 0}).sort("task_id", 1).to_list(None)
    return tasks


@api_router.put("/tasks/{task_id}/progress")
async def update_progress(task_id: int, update: ProgressUpdate):
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    if not task.get("is_leaf"):
        raise HTTPException(400, "Can only update leaf task progress")
    old_progress = task["progress"]
    old_status = task["status"]
    progress = max(0, min(100, round(update.progress, 1)))
    status = determine_status(progress, task)
    await db.tasks.update_one(
        {"task_id": task_id},
        {"$set": {"progress": progress, "status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_history(task_id, "progress_update", "progress", old_progress, progress, update.update_notes or "")
    if old_status != status:
        await log_history(task_id, "status_change", "status", old_status, status, "Auto-derived from progress")
    if task.get("parent_task_id") is not None:
        await rollup_progress(task["parent_task_id"])
    return await db.tasks.find_one({"task_id": task_id}, {"_id": 0})


@api_router.put("/tasks/{task_id}/dates")
async def update_dates(task_id: int, update: DateUpdate):
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    start = date_type.fromisoformat(update.start_date)
    end = date_type.fromisoformat(update.end_date)
    duration = max((end - start).days + 1, 1)
    old_dates = f"{task['start_date']} to {task['end_date']}"
    new_dates = f"{update.start_date} to {update.end_date}"
    await db.tasks.update_one(
        {"task_id": task_id},
        {"$set": {"start_date": update.start_date, "end_date": update.end_date, "duration": duration, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_history(task_id, "date_change", "dates", old_dates, new_dates, update.update_notes or "")
    if task.get("parent_task_id") is not None:
        await rollup_progress(task["parent_task_id"])
    return await db.tasks.find_one({"task_id": task_id}, {"_id": 0})


@api_router.put("/tasks/{task_id}/risk")
async def update_risk(task_id: int, update: RiskUpdate):
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    old_risk = task.get("risk_flagged", False)
    status = task["status"]
    if update.risk_flagged and status not in ["completed"]:
        status = "at_risk"
    elif not update.risk_flagged and status == "at_risk":
        status = determine_status(task["progress"], {**task, "risk_flagged": False})
    await db.tasks.update_one(
        {"task_id": task_id},
        {"$set": {"risk_flagged": update.risk_flagged, "risk_notes": update.risk_notes or "", "status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await log_history(task_id, "risk_change", "risk_flagged", old_risk, update.risk_flagged, update.update_notes or update.risk_notes or "")
    if task.get("parent_task_id") is not None:
        await rollup_progress(task["parent_task_id"])
    return await db.tasks.find_one({"task_id": task_id}, {"_id": 0})


@api_router.put("/tasks/{task_id}/notes")
async def update_notes(task_id: int, update: NotesUpdate):
    task = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    await db.tasks.update_one(
        {"task_id": task_id},
        {"$set": {"notes": update.notes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return await db.tasks.find_one({"task_id": task_id}, {"_id": 0})


@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    all_tasks = await db.tasks.find({}, {"_id": 0}).sort("task_id", 1).to_list(None)
    root_task = next((t for t in all_tasks if t["task_id"] == 1), None)
    overall_progress = root_task["progress"] if root_task else 0

    leaf_tasks = [t for t in all_tasks if t.get("is_leaf") and not t.get("exclude_from_rollup")]
    status_counts = {}
    for t in leaf_tasks:
        s = t["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    phase_map = {
        "pre_construction": "Pre-Construction",
        "admin_academic": "Admin & Academic Block",
        "auditorium": "Auditorium Block",
        "residential": "Residential Quarters & Hostels",
        "external": "External Development"
    }
    phase_stats = []
    for key, name in phase_map.items():
        pt = [t for t in all_tasks if t["phase"] == key]
        pl = [t for t in pt if t.get("is_leaf") and not t.get("exclude_from_rollup")]
        td = sum(t["duration"] for t in pl)
        prog = sum(t["progress"] * t["duration"] for t in pl) / td if td > 0 else 0
        ps = {}
        for t in pl:
            ps[t["status"]] = ps.get(t["status"], 0) + 1
        phase_stats.append({"key": key, "name": name, "progress": round(prog, 1), "total_tasks": len(pt), "leaf_tasks": len(pl), "status_counts": ps})

    at_risk = [{"task_id": t["task_id"], "name": t["name"], "progress": t["progress"], "end_date": t["end_date"], "risk_notes": t.get("risk_notes", "")} for t in all_tasks if t.get("risk_flagged")]
    delayed = [{"task_id": t["task_id"], "name": t["name"], "progress": t["progress"], "end_date": t["end_date"]} for t in all_tasks if t["status"] == "delayed"]

    return {
        "overall_progress": overall_progress,
        "total_tasks": len(all_tasks),
        "leaf_tasks": len(leaf_tasks),
        "status_counts": status_counts,
        "phase_stats": phase_stats,
        "at_risk_items": at_risk,
        "delayed_items": delayed,
        "project_start": "2025-03-28",
        "project_end": "2027-07-31"
    }


@api_router.get("/tasks/{task_id}/history")
async def get_task_history(task_id: int):
    history = await db.task_history.find({"task_id": task_id}, {"_id": 0}).sort("timestamp", -1).to_list(100)
    return history


@api_router.get("/history/recent")
async def get_recent_history(limit: int = Query(30)):
    history = await db.task_history.find({}, {"_id": 0}).sort("timestamp", -1).to_list(min(limit, 100))
    task_ids = list(set(h["task_id"] for h in history))
    tasks = await db.tasks.find({"task_id": {"$in": task_ids}}, {"_id": 0, "task_id": 1, "name": 1}).to_list(None)
    name_map = {t["task_id"]: t["name"] for t in tasks}
    for h in history:
        h["task_name"] = name_map.get(h["task_id"], f"Task #{h['task_id']}")
    return history


@api_router.get("/reports/generate")
async def generate_report(
    format: str = Query("excel"),
    report_type: str = Query("full"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    phase: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    include_history: bool = Query(False)
):
    """
    Generate a report with optional filters.
    report_type: full, daily, weekly, monthly
    start_date/end_date: Filter tasks by date range (YYYY-MM-DD)
    phase: Filter by phase (pre_construction, admin_academic, etc.)
    status: Filter by status (completed, in_progress, delayed, at_risk, not_started)
    include_history: Include recent update history in report
    """
    all_tasks = await db.tasks.find({}, {"_id": 0}).sort("task_id", 1).to_list(None)
    
    # Apply phase filter
    if phase and phase != "all":
        all_tasks = [t for t in all_tasks if t.get("phase") == phase]
    
    # Apply status filter
    if status and status != "all":
        all_tasks = [t for t in all_tasks if t.get("status") == status]
    
    # Apply date range filter (tasks active within the range)
    if start_date or end_date:
        filtered_tasks = []
        for t in all_tasks:
            task_start = t.get("start_date", "1900-01-01")
            task_end = t.get("end_date", "2099-12-31")
            
            # Task overlaps with the requested range
            range_start = start_date or "1900-01-01"
            range_end = end_date or "2099-12-31"
            
            # Check if task period overlaps with filter period
            if task_start <= range_end and task_end >= range_start:
                filtered_tasks.append(t)
        all_tasks = filtered_tasks
    
    # Get history if requested
    history_data = []
    if include_history:
        # Get history for the date range
        history_query = {}
        if start_date:
            history_query["timestamp"] = {"$gte": start_date}
        if end_date:
            if "timestamp" in history_query:
                history_query["timestamp"]["$lte"] = end_date + "T23:59:59"
            else:
                history_query["timestamp"] = {"$lte": end_date + "T23:59:59"}
        
        history_data = await db.task_history.find(history_query, {"_id": 0}).sort("timestamp", -1).to_list(500)
        
        # Add task names to history
        task_ids = list(set(h["task_id"] for h in history_data))
        tasks_map = {t["task_id"]: t["name"] for t in await db.tasks.find({"task_id": {"$in": task_ids}}, {"_id": 0, "task_id": 1, "name": 1}).to_list(None)}
        for h in history_data:
            h["task_name"] = tasks_map.get(h["task_id"], f"Task #{h['task_id']}")
    
    # Generate report metadata
    report_meta = {
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "phase": phase,
        "status": status,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    }
    
    if format == "excel":
        return generate_excel_enhanced(all_tasks, history_data, report_meta)
    return generate_pdf_enhanced(all_tasks, history_data, report_meta)


def generate_excel(tasks):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Summary"

    headers = ["ID", "Task Name", "Level", "Phase", "Status", "Progress (%)", "Start Date", "End Date", "Duration", "Risk", "Risk Notes"]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    sc_map = {"completed": "10B981", "in_progress": "3B82F6", "delayed": "EF4444", "at_risk": "F59E0B", "not_started": "94A3B8"}

    for r, t in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=t["task_id"])
        ws.cell(row=r, column=2, value="  " * t["level"] + t["name"])
        ws.cell(row=r, column=3, value=t["level"])
        ws.cell(row=r, column=4, value=t["phase"].replace("_", " ").title())
        sc = ws.cell(row=r, column=5, value=t["status"].replace("_", " ").title())
        color = sc_map.get(t["status"], "94A3B8")
        sc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sc.font = Font(color="FFFFFF", bold=True)
        ws.cell(row=r, column=6, value=t["progress"])
        ws.cell(row=r, column=7, value=t["start_date"])
        ws.cell(row=r, column=8, value=t["end_date"])
        ws.cell(row=r, column=9, value=t["duration"])
        ws.cell(row=r, column=10, value="YES" if t.get("risk_flagged") else "No")
        ws.cell(row=r, column=11, value=t.get("risk_notes", ""))

    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 55)

    ws2 = wb.create_sheet("At Risk Items")
    rh = ["ID", "Task Name", "Progress (%)", "End Date", "Risk Notes"]
    for c, h in enumerate(rh, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    for r, t in enumerate([t for t in tasks if t.get("risk_flagged")], 2):
        ws2.cell(row=r, column=1, value=t["task_id"])
        ws2.cell(row=r, column=2, value=t["name"])
        ws2.cell(row=r, column=3, value=t["progress"])
        ws2.cell(row=r, column=4, value=t["end_date"])
        ws2.cell(row=r, column=5, value=t.get("risk_notes", ""))

    ws3 = wb.create_sheet("Delayed Items")
    for c, h in enumerate(["ID", "Task Name", "Progress (%)", "End Date"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    for r, t in enumerate([t for t in tasks if t["status"] == "delayed"], 2):
        ws3.cell(row=r, column=1, value=t["task_id"])
        ws3.cell(row=r, column=2, value=t["name"])
        ws3.cell(row=r, column=3, value=t["progress"])
        ws3.cell(row=r, column=4, value=t["end_date"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=IIMC_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"})


def generate_excel_enhanced(tasks, history_data, report_meta):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Summary"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Report Header
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="IIMC Amravati - Project Progress Report")
    title_cell.font = Font(bold=True, size=16, color="0F172A")
    title_cell.alignment = Alignment(horizontal="center")

    # Report Metadata
    row = 3
    report_type_label = {"full": "Full Report", "daily": "Daily Report", "weekly": "Weekly Report", "monthly": "Monthly Report"}
    ws.cell(row=row, column=1, value="Report Type:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=report_type_label.get(report_meta["report_type"], "Full Report"))
    
    if report_meta.get("start_date") or report_meta.get("end_date"):
        row += 1
        ws.cell(row=row, column=1, value="Date Range:").font = Font(bold=True)
        date_range = f"{report_meta.get('start_date', 'Start')} to {report_meta.get('end_date', 'Present')}"
        ws.cell(row=row, column=2, value=date_range)
    
    if report_meta.get("phase") and report_meta.get("phase") != "all":
        row += 1
        ws.cell(row=row, column=1, value="Phase:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_meta["phase"].replace("_", " ").title())
    
    if report_meta.get("status") and report_meta.get("status") != "all":
        row += 1
        ws.cell(row=row, column=1, value="Status:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_meta["status"].replace("_", " ").title())
    
    row += 1
    ws.cell(row=row, column=1, value="Generated:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=report_meta["generated_at"])
    
    row += 1
    ws.cell(row=row, column=1, value="Total Tasks:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=str(len(tasks)))

    # Summary Statistics
    row += 2
    leaf_tasks = [t for t in tasks if t.get("is_leaf") and not t.get("exclude_from_rollup")]
    status_counts = {}
    for t in leaf_tasks:
        s = t["status"]
        status_counts[s] = status_counts.get(s, 0) + 1
    
    summary_headers = ["Total Leaf", "Completed", "In Progress", "Delayed", "At Risk", "Not Started"]
    summary_values = [
        len(leaf_tasks),
        status_counts.get("completed", 0),
        status_counts.get("in_progress", 0),
        status_counts.get("delayed", 0),
        status_counts.get("at_risk", 0),
        status_counts.get("not_started", 0)
    ]
    
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    row += 1
    status_colors = {"Total Leaf": "0F172A", "Completed": "10B981", "In Progress": "3B82F6", "Delayed": "EF4444", "At Risk": "F59E0B", "Not Started": "94A3B8"}
    for c, (h, v) in enumerate(zip(summary_headers, summary_values), 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Task List Header
    row += 2
    headers = ["ID", "Task Name", "Level", "Phase", "Status", "Progress (%)", "Start Date", "End Date", "Duration", "Risk", "Risk Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    sc_map = {"completed": "10B981", "in_progress": "3B82F6", "delayed": "EF4444", "at_risk": "F59E0B", "not_started": "94A3B8"}

    for t in tasks:
        row += 1
        ws.cell(row=row, column=1, value=t["task_id"]).border = thin_border
        ws.cell(row=row, column=2, value="  " * t["level"] + t["name"]).border = thin_border
        ws.cell(row=row, column=3, value=t["level"]).border = thin_border
        ws.cell(row=row, column=4, value=t["phase"].replace("_", " ").title()).border = thin_border
        sc = ws.cell(row=row, column=5, value=t["status"].replace("_", " ").title())
        color = sc_map.get(t["status"], "94A3B8")
        sc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sc.font = Font(color="FFFFFF", bold=True)
        sc.border = thin_border
        ws.cell(row=row, column=6, value=t["progress"]).border = thin_border
        ws.cell(row=row, column=7, value=t["start_date"]).border = thin_border
        ws.cell(row=row, column=8, value=t["end_date"]).border = thin_border
        ws.cell(row=row, column=9, value=t["duration"]).border = thin_border
        ws.cell(row=row, column=10, value="YES" if t.get("risk_flagged") else "No").border = thin_border
        ws.cell(row=row, column=11, value=t.get("risk_notes", "")).border = thin_border

    # Adjust column widths
    for col in ws.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 55)

    # At Risk Items Sheet
    ws2 = wb.create_sheet("At Risk Items")
    rh = ["ID", "Task Name", "Progress (%)", "End Date", "Risk Notes"]
    for c, h in enumerate(rh, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        cell.border = thin_border
    for r, t in enumerate([t for t in tasks if t.get("risk_flagged")], 2):
        ws2.cell(row=r, column=1, value=t["task_id"]).border = thin_border
        ws2.cell(row=r, column=2, value=t["name"]).border = thin_border
        ws2.cell(row=r, column=3, value=t["progress"]).border = thin_border
        ws2.cell(row=r, column=4, value=t["end_date"]).border = thin_border
        ws2.cell(row=r, column=5, value=t.get("risk_notes", "")).border = thin_border

    for col in ws2.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(ml + 2, 55)

    # Delayed Items Sheet
    ws3 = wb.create_sheet("Delayed Items")
    for c, h in enumerate(["ID", "Task Name", "Progress (%)", "End Date"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        cell.border = thin_border
    for r, t in enumerate([t for t in tasks if t["status"] == "delayed"], 2):
        ws3.cell(row=r, column=1, value=t["task_id"]).border = thin_border
        ws3.cell(row=r, column=2, value=t["name"]).border = thin_border
        ws3.cell(row=r, column=3, value=t["progress"]).border = thin_border
        ws3.cell(row=r, column=4, value=t["end_date"]).border = thin_border

    for col in ws3.columns:
        ml = max(len(str(c.value or "")) for c in col)
        ws3.column_dimensions[col[0].column_letter].width = min(ml + 2, 55)

    # Update History Sheet (if requested)
    if history_data:
        ws4 = wb.create_sheet("Update History")
        hist_headers = ["Timestamp", "Task ID", "Task Name", "Action", "Field", "Old Value", "New Value", "Notes"]
        for c, h in enumerate(hist_headers, 1):
            cell = ws4.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            cell.border = thin_border
        
        for r, h in enumerate(history_data, 2):
            ws4.cell(row=r, column=1, value=h.get("timestamp", "")).border = thin_border
            ws4.cell(row=r, column=2, value=h.get("task_id", "")).border = thin_border
            ws4.cell(row=r, column=3, value=h.get("task_name", "")).border = thin_border
            ws4.cell(row=r, column=4, value=h.get("action", "").replace("_", " ").title()).border = thin_border
            ws4.cell(row=r, column=5, value=h.get("field", "")).border = thin_border
            ws4.cell(row=r, column=6, value=h.get("old_value", "")).border = thin_border
            ws4.cell(row=r, column=7, value=h.get("new_value", "")).border = thin_border
            ws4.cell(row=r, column=8, value=h.get("notes", "")).border = thin_border

        for col in ws4.columns:
            ml = max(len(str(c.value or "")) for c in col)
            ws4.column_dimensions[col[0].column_letter].width = min(ml + 2, 55)

    # Generate filename based on report type
    report_type = report_meta.get("report_type", "full")
    date_suffix = datetime.now().strftime('%Y%m%d')
    if report_meta.get("start_date"):
        date_suffix = report_meta["start_date"].replace("-", "")
    
    filename = f"IIMC_{report_type.title()}_Report_{date_suffix}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


def generate_pdf(tasks):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("IIMC Amravati - Project Progress Report", styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    leaf = [t for t in tasks if t.get("is_leaf") and not t.get("exclude_from_rollup")]
    counts = {}
    for t in leaf:
        counts[t["status"]] = counts.get(t["status"], 0) + 1

    sd = [["Total", "Completed", "In Progress", "Delayed", "At Risk", "Not Started"],
          [str(len(leaf)), str(counts.get("completed", 0)), str(counts.get("in_progress", 0)),
           str(counts.get("delayed", 0)), str(counts.get("at_risk", 0)), str(counts.get("not_started", 0))]]
    st = Table(sd, colWidths=[1.5*inch]*6)
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 20))

    td = [["ID", "Task", "Phase", "Status", "Progress", "Start", "End"]]
    for t in tasks:
        td.append([str(t["task_id"]), ("  " * t["level"] + t["name"])[:45], t["phase"].replace("_", " ").title()[:15],
                    t["status"].replace("_", " ").title(), f"{t['progress']}%", t["start_date"], t["end_date"]])

    tt = Table(td, colWidths=[0.4*inch, 3.2*inch, 1.2*inch, 0.9*inch, 0.7*inch, 0.8*inch, 0.8*inch])
    tstyle = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]
    scm = {"completed": '#10B981', "in_progress": '#3B82F6', "delayed": '#EF4444', "at_risk": '#F59E0B', "not_started": '#94A3B8'}
    for i, t in enumerate(tasks, 1):
        c = scm.get(t["status"], '#94A3B8')
        tstyle.append(('BACKGROUND', (3, i), (3, i), colors.HexColor(c)))
        tstyle.append(('TEXTCOLOR', (3, i), (3, i), colors.white))
    tt.setStyle(TableStyle(tstyle))
    elements.append(tt)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=IIMC_Report_{datetime.now().strftime('%Y%m%d')}.pdf"})


def generate_pdf_enhanced(tasks, history_data, report_meta):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("IIMC Amravati - Project Progress Report", styles['Title']))
    
    # Report metadata
    report_type_label = {"full": "Full Report", "daily": "Daily Report", "weekly": "Weekly Report", "monthly": "Monthly Report"}
    meta_text = f"<b>Report Type:</b> {report_type_label.get(report_meta.get('report_type', 'full'), 'Full Report')}"
    
    if report_meta.get("start_date") or report_meta.get("end_date"):
        date_range = f"{report_meta.get('start_date', 'Start')} to {report_meta.get('end_date', 'Present')}"
        meta_text += f" | <b>Period:</b> {date_range}"
    
    if report_meta.get("phase") and report_meta.get("phase") != "all":
        meta_text += f" | <b>Phase:</b> {report_meta['phase'].replace('_', ' ').title()}"
    
    if report_meta.get("status") and report_meta.get("status") != "all":
        meta_text += f" | <b>Status:</b> {report_meta['status'].replace('_', ' ').title()}"
    
    elements.append(Paragraph(meta_text, styles['Normal']))
    elements.append(Paragraph(f"<b>Generated:</b> {report_meta['generated_at']} | <b>Tasks:</b> {len(tasks)}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary Statistics
    leaf = [t for t in tasks if t.get("is_leaf") and not t.get("exclude_from_rollup")]
    counts = {}
    for t in leaf:
        counts[t["status"]] = counts.get(t["status"], 0) + 1

    sd = [["Total Leaf", "Completed", "In Progress", "Delayed", "At Risk", "Not Started"],
          [str(len(leaf)), str(counts.get("completed", 0)), str(counts.get("in_progress", 0)),
           str(counts.get("delayed", 0)), str(counts.get("at_risk", 0)), str(counts.get("not_started", 0))]]
    st = Table(sd, colWidths=[1.5*inch]*6)
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 20))

    # Task List
    if tasks:
        elements.append(Paragraph("<b>Task Details</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        td = [["ID", "Task", "Phase", "Status", "Progress", "Start", "End"]]
        for t in tasks:
            td.append([str(t["task_id"]), ("  " * t["level"] + t["name"])[:45], t["phase"].replace("_", " ").title()[:15],
                        t["status"].replace("_", " ").title(), f"{t['progress']}%", t["start_date"], t["end_date"]])

        tt = Table(td, colWidths=[0.4*inch, 3.2*inch, 1.2*inch, 0.9*inch, 0.7*inch, 0.8*inch, 0.8*inch])
        tstyle = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]
        scm = {"completed": '#10B981', "in_progress": '#3B82F6', "delayed": '#EF4444', "at_risk": '#F59E0B', "not_started": '#94A3B8'}
        for i, t in enumerate(tasks, 1):
            c = scm.get(t["status"], '#94A3B8')
            tstyle.append(('BACKGROUND', (3, i), (3, i), colors.HexColor(c)))
            tstyle.append(('TEXTCOLOR', (3, i), (3, i), colors.white))
        tt.setStyle(TableStyle(tstyle))
        elements.append(tt)

    # At Risk Items
    at_risk = [t for t in tasks if t.get("risk_flagged")]
    if at_risk:
        elements.append(PageBreak())
        elements.append(Paragraph("<b>At Risk Items</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        rd = [["ID", "Task Name", "Progress", "End Date", "Risk Notes"]]
        for t in at_risk:
            rd.append([str(t["task_id"]), t["name"][:40], f"{t['progress']}%", t["end_date"], (t.get("risk_notes", "") or "")[:30]])
        
        rt = Table(rd, colWidths=[0.5*inch, 3*inch, 0.8*inch, 1*inch, 3*inch])
        rt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(rt)

    # Delayed Items
    delayed = [t for t in tasks if t["status"] == "delayed"]
    if delayed:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("<b>Delayed Items</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        dd = [["ID", "Task Name", "Progress", "End Date"]]
        for t in delayed:
            dd.append([str(t["task_id"]), t["name"][:50], f"{t['progress']}%", t["end_date"]])
        
        dt = Table(dd, colWidths=[0.5*inch, 5*inch, 0.8*inch, 1*inch])
        dt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(dt)

    # Update History
    if history_data:
        elements.append(PageBreak())
        elements.append(Paragraph("<b>Update History</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        hd = [["Timestamp", "Task", "Action", "Change", "Notes"]]
        for h in history_data[:50]:  # Limit to 50 entries in PDF
            change = f"{h.get('old_value', '')} → {h.get('new_value', '')}"
            hd.append([
                h.get("timestamp", "")[:16],
                f"#{h.get('task_id', '')} {h.get('task_name', '')[:25]}",
                h.get("action", "").replace("_", " ").title()[:15],
                change[:20],
                (h.get("notes", "") or "")[:25]
            ])
        
        ht = Table(hd, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1.5*inch, 2*inch])
        ht.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        elements.append(ht)

    doc.build(elements)
    buf.seek(0)
    
    # Generate filename
    report_type = report_meta.get("report_type", "full")
    date_suffix = datetime.now().strftime('%Y%m%d')
    if report_meta.get("start_date"):
        date_suffix = report_meta["start_date"].replace("-", "")
    
    filename = f"IIMC_{report_type.title()}_Report_{date_suffix}.pdf"
    
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@api_router.post("/tasks/reseed")
async def reseed_tasks():
    await db.tasks.drop()
    from seed_data import get_seed_tasks
    tasks = get_seed_tasks()
    await db.tasks.insert_many(tasks)
    await db.tasks.create_index("task_id", unique=True)
    await db.tasks.create_index("parent_task_id")
    await db.tasks.create_index("phase")
    return {"message": f"Reseeded {len(tasks)} tasks"}


app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True,
                   allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
                   allow_methods=["*"], allow_headers=["*"])

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
